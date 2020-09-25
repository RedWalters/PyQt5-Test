from sqlalchemy import create_engine, event
import sqlalchemy.sql.default_comparator
from multiprocessing import Pool, freeze_support
from import_file import import_file
from functools import partial
from ntpath import basename
import sqlalchemy as sa
from io import StringIO
import psycopg2 as pg
import pandas as pd
import itertools
import pyodbc
import glob
import time
import csv
import sys
import os
import xlrd
from openpyxl import load_workbook

#wb = load_workbook('sample.xlsx')
#ws = wb['sample']
# 
#data = ws.values
#Get the first line in file as a header line
#columns = next(data)[0:]
#Create a DataFrame based on the second and subsequent lines of data
#df = pd.DataFrame(data, columns=columns)

#impModule = import_file(formatting)

def read_excel(filename):#, formatting):
    
    #pd.set_option('display.max_colwidth', -1)
    
    #dfs = []
    
    #for i in formatting:
        #dfs.append(pd.read_excel(filename, sheet_name = formatting))
    
    #return dfs
    #xls = xlrd.open_workbook(filename, on_demand=0)
    #print(xls.sheet_names())
    
    wb= load_workbook(filename, read_only=True)    
    
    sheets = wb.sheetnames
    
    sheet = wb[sheets[0]]
    
    data = sheet.values
    
    columns = next(data)[0:]
    # Create a DataFrame based on the second and subsequent lines of data
    df = pd.DataFrame(data, columns=columns, index = None)
    #columns = next(data)[0:]
    
    #dfs = []
    
    #impModule = import_file(formatting)
    
    #for i in wb.sheetnames:
        #data = wb[i].values
        #df = pd.DataFrame(data).fillna('')
        #df['File Name'] = basename(filename)
        #df['Sheet Name'] = i
        #dfs.append(df)
        
    #with Pool(4) as pool:
     #   impModule = import_file(formatting)
      #  df_list = pool.map(impModule.formatting, dfs)
    
    #print(filename)
    #print(df_list)
    #f = pd.DataFrame(data, columns=columns)
    
    #file = pd.ExcelFile(filename, on_demand = 0)
    #sheets = file.sheet_names
    #dfs = []
    #df = pd.DataFrame()
    #for i in sheets:
        #print(f'     Sheet: ' + i)
        #df = file.parse(i, header = None, encoding="utf-8-sig", ignore_index = True).fillna('')
        #df['File Name'] = basename(filename)
        #df['Sheet Name'] = i
        #dfs.append(impModule.formatting(df))
    
    #print(dfs)
    #for key, value in test.items():
        #print(key)
        #print(value)
    
    #df = impModule.formatting(dfs)
    
    return df
def pandasread_excel(filename):
    
    xl = pd.ExcelFile(filename)
    
    sheetnames = xl.sheet_names
    
    sheet = xl.parse(sheetnames[0])
    
    return sheet


if __name__ == '__main__':    

    file = 'K:\\A & A\\Cardiff\\Audit\\Clients\\Open\\S\\Spotlight\\2. Staff Folders\\JWalters\\__Python\\Test Data and VBA\\Baybridge Housing\\Oct 2018 GL details - Yardi Trust.xlsx'
    frmat = 'K:\\A & A\\Cardiff\\Audit\\Clients\\Open\\S\\Spotlight\\2. Staff Folders\\JWalters\\__Python\\Formatting Scripts\\YardiTrust.py'

    s = time.time()
    
    #impModule = import_file(frmat)   
    print(read_excel(file))
    
    print(time.time() - s)
    
    
    s = time.time()
    
    print(pandasread_excel(file))
    
    print(time.time() - s)
    
    #files = []
    
    #files.append(file)
    
    #wb= load_workbook(file, read_only=True)   
    #with Pool(4) as pool:        
         
        #pool.map(partial(read_excel, formatting = frmat),  files)
        
        #pool.map(impModule.formatting, read_excel(file))
    
    #read_excel(file, frmat)


    





