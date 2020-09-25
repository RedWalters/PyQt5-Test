from PyQt5.QtWidgets import QApplication, QWidget, QLineEdit, QFileDialog,QHBoxLayout,  QPushButton, QRadioButton, QGridLayout, QLabel, QVBoxLayout, QTextEdit, QProgressBar, QComboBox, QCheckBox
from sqlalchemy import create_engine, event
import sqlalchemy.sql.default_comparator
from PyQt5.QtCore import QBasicTimer, QThread, pyqtSignal, Qt
from multiprocessing import Pool, freeze_support
from PyQt5.QtGui import QTextCursor
from urllib.parse import quote_plus
from shutil import move, copy
from functools import partial
from ntpath import basename
import sqlalchemy as sa
import pandas as pd
import pyodbc
import glob
import time
import sys
import os
import itertools
from import_file import import_file
from openpyxl import load_workbook
from pandas.io import sql
import gc

#from dfExcel import make_df_from_excel, make_df_from_excelDepen, methodSelection, saveToCsv, write_df_to_sql
     
dfExcel = import_file('K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/Exe/dfExcel.py')

  
def left(s, amount):
    return s[:amount]

def right(s, amount):
    return s[-amount:]

def mid(s, offset, amount):
    return s[offset:offset+amount]
    
freeze_support()

class App(QWidget):
    
    def __init__(self):
        super().__init__()

        self.title = 'Excel Formatter'
        self.left = 30
        self.top = 50
        self.width = 640
        self.height = 480
        self.initUI()
        
        
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)        
        self.setFixedSize(800, 550)
        
        browse = QPushButton('Browse: ', self)
        browse.setToolTip('Browse for files to be Formatted: ')
        browse.clicked.connect(lambda: self.browse_on_click(App))
        browse.setMinimumHeight(25)
        browse.setMaximumHeight(25)
        browse.setMinimumWidth(135)
        browse.setMaximumWidth(135)
        
        self.browseFunc = QPushButton('Import Script: ', self)
        self.browseFunc.setToolTip('Import Bespoke Python Script: ')
        self.browseFunc.clicked.connect(lambda: self.importScript(App))
        self.browseFunc.setMinimumHeight(25)
        self.browseFunc.setMaximumHeight(25)
        self.browseFunc.setMinimumWidth(135)
        self.browseFunc.setMaximumWidth(135)
      
        self.saveTo = QPushButton('Save to: ', self)
        self.saveTo.setToolTip('Select location to save output: ')
        self.saveTo.clicked.connect(lambda: self.saveTo_on_click(App))
        self.saveTo.setMinimumHeight(25)
        self.saveTo.setMaximumHeight(25)
        self.saveTo.setMinimumWidth(135)
        self.saveTo.setMaximumWidth(135)
        self.saveTo.setEnabled(False)
        
        self.run = QPushButton('Run: ', self)
        self.run.setToolTip('Run Program: ')
        self.run.clicked.connect(self.run_on_click)
        self.run.setMinimumHeight(25)
        self.run.setMaximumHeight(25)
        self.run.setMinimumWidth(135)
        self.run.setMaximumWidth(135)
        
        reset = QPushButton('Reset View', self)
        reset.setToolTip('Reset UI Elements: ')
        reset.clicked.connect(self.reset_on_click)
        reset.setMinimumHeight(25)
        reset.setMaximumHeight(25)
        reset.setMinimumWidth(135)
        reset.setMaximumWidth(135)        
        
        self.sheets = QCheckBox('Cont. Sheets?')
        self.sheets.setChecked(False)
        self.combine = QCheckBox('Combine Files?')
        self.combine.setChecked(True)
        self.combine.setEnabled(True)
        self.zip = QCheckBox('Zip Files?')
        self.zip.hide()
        self.zip.setChecked(False)
        self.zip.setEnabled(False)
        self.zip.toggled.connect(self.onClicked)
        self.noForm = QCheckBox('No Formatting?')
        self.noForm.setChecked(False)
        self.noForm.toggled.connect(self.noFormOnClicked) 
        
        self.delimIn = QComboBox()
        self.delimIn.addItems(['Input CSV Delim: ',',','|','@',';'])
        self.delimIn.setEditable(True)
        self.delimIn.lineEdit().setAlignment(Qt.AlignCenter)
        self.delimIn.lineEdit().setReadOnly(True)
        self.delimIn.setMinimumHeight(25)
        self.delimIn.setMaximumHeight(25)
        self.delimIn.setMinimumWidth(135)
        self.delimIn.setMaximumWidth(135)    
        
        self.delimOut = QComboBox()
        self.delimOut.addItems(['Output CSV Delim: ',',','|','@',';'])
        self.delimOut.setEditable(True)
        self.delimOut.lineEdit().setAlignment(Qt.AlignCenter)
        self.delimOut.lineEdit().setReadOnly(True)
        self.delimOut.setMinimumHeight(25)
        self.delimOut.setMaximumHeight(25)
        self.delimOut.setMinimumWidth(135)
        self.delimOut.setMaximumWidth(135) 
        
        
        self.threads = QComboBox()
        self.threads.addItems(['No of Threads: ','1', '2', '3', '4', '6', '8', '12', '16', '20']) 
        self.threads.setEditable(True)
        self.threads.lineEdit().setAlignment(Qt.AlignCenter)
        self.threads.lineEdit().setReadOnly(True)
        self.threads.setMinimumHeight(25)
        self.threads.setMaximumHeight(25)
        self.threads.setMinimumWidth(135)
        self.threads.setMaximumWidth(135)        
        
        self.files = []
        self.saveFile = ''
        self.imprt = ''
        
        self.financialSystem  = QComboBox()
        self.financialSystem.addItems(['Financial System:', 'No Formatting', 'Sage200 - ACr'])
        self.financialSystem.setEditable(True)
        self.financialSystem.lineEdit().setAlignment(Qt.AlignCenter)
        self.financialSystem.lineEdit().setReadOnly(True)
        self.financialSystem.setMinimumHeight(25)
        self.financialSystem.setMaximumHeight(25)
        self.financialSystem.setMinimumWidth(135)
        self.financialSystem.setMaximumWidth(135)   
        
        self.serverName = 'UKDC250016'
        
        self.engine = sa.create_engine('mssql+pyodbc://'+self.serverName+'/master?driver=SQL+Server')

        conn = self.engine.connect()
        rows = conn.execute("select name FROM sys.databases;")
        dbs= []
        for row in rows:
            dbs.append(row["name"])        
        
        self.server = QComboBox()
        self.server.addItem('Server: ', dbs) 
        self.server.setEditable(True)
        self.server.lineEdit().setAlignment(Qt.AlignCenter)
        self.server.lineEdit().setReadOnly(True)
        self.server.setMinimumHeight(25)
        self.server.setMaximumHeight(25)
        self.server.setMinimumWidth(180)
        self.server.setMaximumWidth(180)
        
        self.server.addItem('UKDC250016', dbs)
        self.server.addItem('UKDC250016\SQL2016', dbs)        
        self.server.addItem('UKDC110029', dbs)
        self.server.addItem('UKDC210033', dbs)
        self.server.addItem('UKDC210033\SQL2016', dbs)
        self.server.addItem('UKDC150016', dbs)
        self.server.addItem('UKDC150016\SQL2016', dbs)
        self.server.addItem('UKDC103023', dbs)
        self.server.addItem('UKDC203023', dbs) 
        self.server.currentIndexChanged.connect(self.indexChanged)        
        
        self.database = QComboBox()
        self.database.addItem('Database: ', dbs) 
        self.database.setEditable(True)
        self.database.lineEdit().setAlignment(Qt.AlignCenter)
        self.database.lineEdit().setReadOnly(True)
        self.database.setMinimumHeight(25)
        self.database.setMaximumHeight(25)
        self.database.setMinimumWidth(286)
        self.database.setMaximumWidth(286)   
        
        self.step = 0
        self.timer = QBasicTimer()
        
        self.linePathBrowse = QLineEdit(self)
        self.linePathSave = QLineEdit(self)
        
        self.output = QTextEdit()
        self.output.setReadOnly(True)
        
        self.pbar = QProgressBar(self)
        self.pbar.setTextVisible(False)          
 
        self.grid = QGridLayout()
        self.grid.setSpacing(10)
        
        self.setLayout(self.grid)   
        
        self.process = QTextEdit()
        self.process.moveCursor(QTextCursor.Start)
        self.process.ensureCursorVisible()
        self.process.setLineWrapColumnOrWidth(128)
        self.process.setLineWrapMode(QTextEdit.FixedPixelWidth)
        self.process.setMinimumWidth(135)
        self.process.setMaximumWidth(135) 
        
        self.process2 = QTextEdit()
        self.process2.moveCursor(QTextCursor.Start)
        self.process2.ensureCursorVisible()
        self.process2.setLineWrapColumnOrWidth(128)
        self.process2.setLineWrapMode(QTextEdit.FixedPixelWidth)
        self.process2.setMinimumWidth(135)
        self.process2.setMaximumWidth(135) 
        
        self.sql   = QCheckBox('To DB')
        self.sql.setChecked(True)
        self.sql.toggled.connect(self.SQLOnClicked)
        self.sql.setMinimumWidth(65)
        self.sql.setMaximumWidth(65)  
        
        self.csv   = QCheckBox('To CSV')
        self.csv.toggled.connect(self.SQLOnClicked)
        self.csv.setMinimumWidth(67)
        self.csv.setMaximumWidth(67)
        
        sublayout1 = QHBoxLayout()
        sublayout1.addWidget(self.csv)
        sublayout1.addWidget(self.sql)
        
        self.tableName = QPushButton('Table Name:')
        self.tableName.setMinimumHeight(25)
        self.tableName.setMaximumHeight(25)
        self.tableName.setMinimumWidth(135)
        self.tableName.setMaximumWidth(135)   

        self.grid.addWidget(browse, 1, 1)
        self.grid.addWidget(self.linePathBrowse, 1, 2, 1, 3)
        self.grid.addWidget(self.saveTo, 3, 1)
        self.saveTo.hide()
        self.grid.addWidget(self.tableName, 3, 1)        
        self.grid.addWidget(self.linePathSave, 3, 2, 1, 3)
        self.grid.addWidget(self.sheets, 5, 1)
        self.grid.addWidget(self.combine, 6, 1)
        self.grid.addWidget(self.threads, 12, 5)
        self.grid.addWidget(self.zip, 7, 1)
        self.grid.addWidget(self.delimIn, 7, 1)
        self.grid.addWidget(self.delimOut, 8, 1)
        self.grid.addWidget(reset, 10, 1)
        self.grid.addWidget(self.run, 13, 5) 
        self.grid.addLayout(sublayout1, 2, 1)
        self.grid.addWidget(self.server, 2, 2)
        self.grid.addWidget(self.database, 2, 3, 1, 2)
        self.grid.addWidget(self.financialSystem, 3, 5)
        self.grid.addWidget(self.noForm, 1, 5)
        self.grid.addWidget(self.browseFunc, 2, 5)
        self.grid.addWidget(self.output, 4, 2, 9, 3)
        self.grid.addWidget(self.pbar, 13, 1, 1, 4) 
        self.spacer= QLabel()
        self.grid.addWidget(self.spacer, 5, 1)
        self.grid.addWidget(self.spacer, 9, 1)
        self.grid.addWidget(self.process, 11, 1, 2, 1)
        self.grid.addWidget(self.process2, 4, 5, 8, 1)
        
        self.show()
        
    def runControl(self):
        if self.run.isEnabled() == True:
            self.run.setEnabled(False)
        else:
            self.run.setEnabled(True)
        
    def indexChanged(self, index):
        if index == 0:
            self.database.clear()
            self.database.addItem('Database: ')
        else:
            self.database.clear()
            data = self.server.itemData(index)
            self.serverName = str(self.server.currentText())
            self.engine = sa.create_engine('mssql+pyodbc://'+self.serverName+'/master?driver=SQL+Server')
            conn = self.engine.connect()
            rows = conn.execute("select name FROM sys.databases;")
            dbs= []
            for row in rows:
                dbs.append(row["name"])
            if data is not None:
                self.database.addItems(dbs)
        
    def noFormOnClicked(self):
        if self.noForm.isChecked():
            self.financialSystem.setEnabled(False)
            self.browseFunc.setEnabled(False)
        else:
            self.financialSystem.setEnabled(True)
            self.browseFunc.setEnabled(True)            
            
    def SQLOnClicked(self):
        if self.csv.isChecked() and self.sql.isChecked():
            self.server.setEnabled(True)
            self.database.setEnabled(True)
            self.saveTo.setEnabled(True)
            self.combine.setEnabled(True)
            self.zip.setEnabled(True)
            self.saveTo.show()
            self.tableName.hide()
            self.linePathSave.setText('')
            self.linePathSave.setReadOnly(True)
        elif self.csv.isChecked():
            self.server.setEnabled(False)
            self.database.setEnabled(False)
            self.saveTo.setEnabled(True)
            self.combine.setEnabled(True)
            self.zip.setEnabled(True)
            self.saveTo.show()
            self.tableName.hide()
            self.linePathSave.setText('')
            self.linePathSave.setReadOnly(True)
        else:
            self.server.setEnabled(True)
            self.database.setEnabled(True)
            self.saveTo.setEnabled(False)
            self.combine.setEnabled(True)
            self.zip.setEnabled(False)
            self.saveTo.hide()
            self.tableName.show()
            self.linePathSave.setText('')
            self.linePathSave.setReadOnly(False)
            
    def onClicked(self):
        if self.zip.isChecked():
            self.output.clear()
        else:
            self.output.append('Highly Recommended for large data sets!')        
        
    def reset_on_click(self):
        self.linePathBrowse.setText('')
        self.linePathSave.setText('')
        self.browseFunc.setText('Import Script: ') 
        self.process.setText('')
        self.sheets.setChecked(False)
        self.combine.setChecked(True)
        self.zip.setChecked(True)
        self.sql.setChecked(True)
        self.csv.setChecked(False)
        self.combine.setEnabled(True)
        self.zip.setEnabled(False)
        self.saveTo.hide()
        self.tableName.show()
        self.noForm.setChecked(False)
        self.browseFunc.setEnabled(True)
        self.financialSystem.setEnabled(True)
        self.server.setCurrentIndex(0)
        self.database.setCurrentIndex(0)
        self.financialSystem.setCurrentIndex(0)
        self.threads.setCurrentIndex(0)
        self.output.clear()
        self.process2.setText('')
    
    def outputToUI(self, text):
        self.output.append(text)
        
    def setTimer(self, timerValue):
        self.pbar.setValue(timerValue)
        
    def importScript(self, App):
        files = self.importDialog()
        return files
    
    def importDialog(self):
        options = QFileDialog.Options()
        #options |= QFileDialog.DontUseNativeDialog
        imprt, _ = QFileDialog.getOpenFileNames(self, "Open Formatting Script: ", "K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/Formatting Scripts", "Python Script (*.py)", options = options)
        if basename('; '.join(imprt)) == '': 
            self.browseFunc.setText('Import Script: ')
        else:
            self.browseFunc.setText(basename('; '.join(imprt))) 
        self.imprt = imprt
            
    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        #options |= QFileDialog.DontUseNativeDialog
        files, _ = QFileDialog.getOpenFileNames(self, "Open Files to be Formatted: ", "K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/Test Data and VBA", "All Files (*);;Excel Workbook (*.xls);;Excel Workbook (*.xlsx);; Excel Macro-Enabled Workbook (*.xlsm)", options = options)
        self.linePathBrowse.setText('; '.join(files)) 
        self.files = files
        
    def saveFileDialog(self):
        options = QFileDialog.Option()
        #options |= QFileDialog.DontUseNativeDialog
        saveFile, _ = QFileDialog.getSaveFileName(self,"Save Output As: ","K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/Test Data and VBA/Formatted.csv","All Files (*);;CSV Files (*.csv)", options=options)
        self.linePathSave.setText(left(saveFile, len(saveFile) -4)) 
        self.saveFile = saveFile
            
    def browse_on_click(self, App):
        files = self.openFileNamesDialog()
        return files
    
    def saveTo_on_click(self, App):
        saveFile = self.saveFileDialog()
        return saveFile
    
    def run_on_click(self, files):
        if self.delimIn.currentText() == 'Input CSV Delim: ':
            delimIn = ','
        else:
            delimIn = self.delimIn.currentText()
        if self.delimOut.currentText() == 'Output CSV Delim: ':
            delimOut = ','
        else:
            delimOut = self.delimOut.currentText()
        if self.csv.isChecked() == True and self.sql.isChecked() == False:
            server = ''
            database = ''
        else:
            server = self.server.currentText()
            database = self.database.currentText()
            
        if self.csv.isChecked() == False:
            saveFileTemp = ''
        else:
            saveFileTemp = self.linePathSave.text()
        
        try:
            if self.files == []:
                self.output.append('No Source File(s) Selected.')
            elif self.saveFile == '' and self.csv.isChecked():
                self.output.append('No Destination Selected.')
            elif self.sql.isChecked() and self.database.currentText() == 'Database: ':
                self.output.append('Please Select Destination Server and Database.')
            elif self.sql.isChecked() and self.linePathSave.text() == '':
                self.output.append('Please Enter Table Name:')
            elif self.noForm.isChecked() == False and ((self.imprt == '' or self.imprt == []) and (self.financialSystem.currentText() == 'Financial System:' or self.financialSystem.currentText() == '')):  
                self.output.append('Please Import Formatting Script: ')
            elif self.csv.isChecked() == False and self.sql.isChecked() == False:
                self.output.append('Please Select Output (SQL/CSV)')
            elif self.financialSystem.currentText() != 'Financial System:' and self.financialSystem.currentText() != '':
                self.imprt = self.financialSystem.currentText()
                self.output.clear()
                if self.threads.currentText() == 'No of Threads: ':
                    self.output.append('No Thread Count Selected:')
                    self.output.append('     Defaulting to 4 Threads.')
                    self.output.append('')
                else:
                    pass
                if os.path.exists(basename('; '.join(self.imprt))) == True:
                    pass
                else:
                    copy('K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/python/System Scripts/' + basename(self.imprt) +'.py', '.')
                impModule = 'K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/python/System Scripts/' + basename(self.imprt) +'.py'
                self.workerThread = WorkerThread(files = self.files, 
                                                 saveFile = saveFileTemp, 
                                                 system = impModule,
                                                 filename = self.imprt,  
                                                 server = server,
                                                 database = database,
                                                 sheets = self.sheets.isChecked(), 
                                                 threads = self.threads.currentText(),
                                                 combine = self.combine.isChecked(),
                                                 zipped = self.zip.isChecked(),
                                                 noForm = self.noForm.isChecked(),
                                                 tableName = self.linePathSave.text(),
                                                 CSVDelimIn = delimIn,
                                                 CSVDelimOut = delimOut) 
                self.runControl()
                self.workerThread.value.connect(self.outputToUI)
                self.workerThread.timerValue.connect(self.setTimer)
                self.workerThread.startValue.connect(self.startTimer)
                self.workerThread.endValue.connect(self.stopTimer)
                self.workerThread.runCont.connect(self.runControl)
                self.workerThread.start()
            else:
                self.output.clear()
                if self.threads.currentText() == 'No of Threads: ':
                    self.output.append('No Thread Count Selected:')
                    self.output.append('     Defaulting to 4 Threads.')
                    self.output.append('')
                else:
                    pass
                try:
                    os.remove(basename('; '.join(self.imprt)))
                except:
                    pass
                if os.path.exists(basename('; '.join(self.imprt))) == True:
                    pass
                elif self.noForm.isChecked() == False:
                    copy('; '.join(self.imprt), '.') 
                    impModule = ('; '.join(self.imprt))
                else:
                    impModule = ''                                
                
                self.workerThread = WorkerThread(files = self.files, 
                                                 saveFile = saveFileTemp, 
                                                 system = impModule,
                                                 filename = self.imprt,  
                                                 server = server,
                                                 database = database,
                                                 sheets = self.sheets.isChecked(), 
                                                 threads = self.threads.currentText(),
                                                 combine = self.combine.isChecked(),
                                                 zipped = self.zip.isChecked(),
                                                 noForm = self.noForm.isChecked(),
                                                 tableName = self.linePathSave.text(),
                                                 CSVDelimIn = delimIn,
                                                 CSVDelimOut = delimOut)
                self.runControl()
                self.workerThread.value.connect(self.outputToUI)
                self.workerThread.timerValue.connect(self.setTimer)
                self.workerThread.startValue.connect(self.startTimer)
                self.workerThread.endValue.connect(self.stopTimer)                
                self.workerThread.runCont.connect(self.runControl)
                self.workerThread.start()
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print(e)
            
    def startTimer(self,startValue):
        self.pbar.setMaximum(startValue)
        self.timer.start(startValue, self)
            
    def timerEvent(self, tevent):
        value = self.pbar.value()
        if value < 35 :
            value += 1
            self.pbar.setValue(value)
        else:
            self.pbar.setValue(0)
    
    def stopTimer(self, inp):
        if inp == 1:
            self.timer.stop()

class WorkerThread(QThread):
    
    def __init__(self, files, server = '', database = '', saveFile = '', system = '', filename = '', sheets = False, threads = 4, combine = True, zipped = True, noForm = False, tableName = '', CSVDelimIn = ',', CSVDelimOut = ','):
        super().__init__()     
        self.files = files
        self.saveFile = saveFile
        self.system = system
        self.filename = filename
        self.sheets = sheets
        if threads == 'No of Threads: ':
            self.threads = 4
        else:
            self.threads = threads
        self.combine = combine
        self.zipped = zipped
        self.noForm = noForm
        self.server = server
        self.database = database
        self.tableName = tableName
        self.delimIn = CSVDelimIn
        self.delimOut = CSVDelimOut
        
        if self.database != '':
            conn =  "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
            quoted = quote_plus(conn)
            new_con = 'mssql+pyodbc:///?odbc_connect={}'.format(quoted)
            self.engine = create_engine(new_con)
            
            @event.listens_for(self.engine, 'before_cursor_execute')
            def receive_before_cursor_execute(conn, cursor, statement, params, context, executemany):
                if executemany:
                    cursor.fast_executemany = True
        else: 
            pass
    
    value = pyqtSignal(str)
    timerValue = pyqtSignal(int)
    startValue = pyqtSignal(int)
    endValue = pyqtSignal(int)
    runCont = pyqtSignal()
        
    def run(self):     
        
        gc.collect()
        flat_files =[]
        non_flat = []
        xls = []
        count = 0
        sheets = {}
        
        try:
            
            dfExcel = import_file('K:/A & A/Cardiff/Audit/Clients/Open/S/Spotlight/2. Staff Folders/JWalters/__Python/Exe/dfExcel.py')
            
            start = time.time()              
                   
            file_list = [filename for filename in self.files]
            
            self.value.emit("Importing Files:")
            
            for i in file_list:               
                self.value.emit("     " + str(basename(i)))
            
            self.startValue.emit(35)
            
            if int(self.threads) == 1:
                            
                formatTime = time.time()
                
                flat_files =[]
                non_flat = []
                xls = []
                count = 0
                
                #ADD DELIM SELECTION USING SELF.DELIM

                for i in file_list:
                    if os.path.splitext(i)[1].lower() == '.csv' or os.path.splitext(i)[1].lower() == '.txt':
                        flat_files.append(i)
                    elif os.path.splitext(i)[1].lower() == '.xls':
                        xls.append(i)
                    else:
                        non_flat.append(i)
                        
                sheets = {}
                for i in non_flat:
                    wb= load_workbook(i, read_only=True)    
                    for j in wb.sheetnames:
                        sheets[count] = [i,j]
                        count +=1
                            
                if self.noForm == True:              
                
                    for i in sheets.items():
                        dfExcel.make_df_from_excelNoForm(i)
                    for i in flat_files:  
                        dfExcel.make_df_from_csvNoForm(i, delim = self.delimIn)
                    for i in xls:
                        dfExcel.make_df_from_excelXLSNoForm(i)
                        
                elif self.sheets == True:
                    for i in non_flat:
                        dfExcel.make_df_from_excelDepen(i, formatting = self.system)
                    for i in xls:
                        dfExcel.make_df_from_excelXLSDepen(i, formatting = self.system)
                    for i in flat_files:  
                        dfExcel.make_df_from_csv(i, formatting = self.system, delim = self.delimIn)
                            
                else:
                    for i in sheets.items():
                        dfExcel.make_df_from_excel(i, formatting = self.system)
                    for i in xls:
                        dfExcel.make_df_from_excelXLS(i, formatting = self.system)
                    for i in flat_files:  
                        dfExcel.make_df_from_csv(i, formatting = self.system, delim = self.delimIn)
                
                
                self.timerValue.emit(35)
                self.endValue.emit(1)                
                        
                self.value.emit('\nRead and Format Time: ' + str(round(time.time() - formatTime,2)))
                     
                try:
                    os.remove(basename('; '.join(self.filename)))
                except:
                    pass
                writestart = time.time()
                if self.database != '' and self.saveFile != '':
                    
                    if '.' in basename(self.saveFile):
                        fileName = os.path.splitext(self.saveFile)[0] 
                        filePath = left(fileName, len(fileName) - len(basename(fileName)))
                    else: 
                        fileName = self.saveFile
                        filePath = left(fileName, len(fileName) - len(basename(fileName)))
                    
                    if self.combine == True:
                        if self.zipped == True:
                            self.value.emit('\nSaving File to csv.gz...')
                            dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                        else:
                            self.value.emit('\nSaving File to csv...')
                            dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                            
                        self.value.emit('\nSaving File to Database...')
                        engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                        dfExcel.uploadCSV(basename(fileName), engine, 50000)
                    
                    else:                      
                                
                        self.value.emit('\nSaving File to Database...')
                        engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                        dfExcel.uploadCSV('pass', engine, 50000)                        
                        
                        if self.zipped == True:
                            csv_files = [i for i in glob.glob('*.{}'.format('gz'))]
                            for i in csv_files:
                                move(i, filePath + i)
                        else:
                            csv_files = [i for i in glob.glob('*.{}'.format('csv'))]
                            for i in csv_files:
                                move(i, filePath + i)  
                    
                                
                elif self.database != '':
                    
                    if '.' in basename(self.tableName):
                        fileName = os.path.splitext(self.tableName)[0] 
                        filePath = left(fileName, len(fileName) - len(basename(fileName)))
                    else: 
                        fileName = self.tableName
                        filePath = left(fileName, len(fileName) - len(basename(fileName)))
                    
                    if self.combine == True:
                        self.value.emit('\nSaving File to Database...')
                        engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                        dfExcel.uploadCSV(basename(fileName), engine, 50000)
                    else:
                        self.value.emit('\nSaving File to Database...')
                        engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                        dfExcel.uploadCSV('pass', engine, 50000)
                    
                
                else: 
                    
                    if '.' in basename(self.saveFile):
                        fileName = os.path.splitext(self.saveFile)[0] 
                        filePath = left(fileName, len(fileName) - len(basename(fileName)))
                    else: 
                        fileName = self.saveFile
                        filePath = left(fileName, len(fileName) - len(basename(fileName)))
                    
                    if self.combine == True:
                        
                        if self.zipped == True:
                            self.value.emit('\nSaving File to csv.gz...')
                            dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                        else:
                            self.value.emit('\nSaving File to csv...')
                            dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                        
                    
                    else:
                        if self.zipped == True:
                            csv_files = [i for i in glob.glob('*.{}'.format('gz'))]
                            for i in csv_files:
                                move(i, filePath + i)
                        else:
                            csv_files = [i for i in glob.glob('*.{}'.format('csv'))]
                            for i in csv_files:
                                move(i, filePath + i)
           
            else:
                with Pool(int(self.threads), maxtasksperchild = 12) as pool:
                            
                    formatTime = time.time()
                            
                    flat_files =[]
                    non_flat = []
                    xls = []
                    count = 0

                    for i in file_list:
                        if os.path.splitext(i)[1].lower() == '.csv' or os.path.splitext(i)[1].lower() == '.txt':
                            flat_files.append(i)
                        elif os.path.splitext(i)[1].lower() == '.xls' :
                            xls.append(i)
                        else:
                            non_flat.append(i)
                      
                    sheets = {}
                    for i in non_flat:
                        wb= load_workbook(i, read_only=True)    
                        for j in wb.sheetnames:
                            sheets[count] = [i,j]
                            count +=1
                        
                    if self.noForm == True:                        
                        flatCount = len(sheets) + 1
                        for i in flat_files:
                            sheets[flatCount] = [i, 'FLAT']
                            flatCount +=1
                        for i in xls:
                            sheets[flatCount] = [i, 'XLS']
                            flatCount +=1
                        
                        pool.map(partial(dfExcel.methodSelection, delim = self.delimIn), sheets.items())
                        
                                
                    elif self.sheets == True:
                        pool.map(partial(dfExcel.make_df_from_excelDepen, formatting = self.system), non_flat)
                        pool.map(partial(dfExcel.make_df_from_excelXLSDepen, formatting = self.system), xls)
                        pool.map(partial(dfExcel.make_df_from_csv, formatting = self.system, delim = self.delimIn), flat_files)
                    else:
                        pool.map(partial(dfExcel.make_df_from_excel, formatting = self.system), sheets.items())
                        pool.map(partial(dfExcel.make_df_from_excelXLS, formatting = self.system), xls)
                        pool.map(partial(dfExcel.make_df_from_csv, formatting = self.system, delim = self.delimIn), flat_files)
                    
                    self.timerValue.emit(35)
                    self.endValue.emit(1)                    
                    
                    self.value.emit('\nRead and Format Time: ' + str(round(time.time() - formatTime,2)))
                                   
                    try:
                        os.remove(basename('; '.join(self.filename)))
                    except:
                        pass
                    writestart = time.time()
                    if self.database != '' and self.saveFile != '':
                                    
                        
                        if '.' in basename(self.saveFile):
                            fileName = os.path.splitext(self.saveFile)[0] 
                            filePath = left(fileName, len(fileName) - len(basename(fileName)))
                        else: 
                            fileName = self.saveFile
                            filePath = left(fileName, len(fileName) - len(basename(fileName)))
                        
                        
                        if self.combine == True:

                            if self.zipped == True:
                                self.value.emit('\nSaving File to csv.gz...')
                                dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)

                            else:
                                self.value.emit('\nSaving File to csv...')
                                dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                               
                            self.value.emit('\nSaving File to Database...')    
                            engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                            dfExcel.uploadCSV(basename(fileName), engine, 50000)
                        
                        else:
                            
                            self.value.emit('\nSaving File to Database...')
                            engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                            dfExcel.uploadCSV('pass', engine, 50000)
                            
                            if self.zipped == True:
                                csv_files = [i for i in glob.glob('*.{}'.format('gz'))]
                                for i in csv_files:
                                    move(i, filePath + i)
                            else:
                                csv_files = [i for i in glob.glob('*.{}'.format('csv'))]
                                for i in csv_files:
                                    move(i, filePath + i)
         
                    elif self.database != '':
                        
                        if '.' in basename(self.tableName):
                            fileName = os.path.splitext(self.tableName)[0] 
                            filePath = left(fileName, len(fileName) - len(basename(fileName)))
                        else: 
                            fileName = self.tableName
                            filePath = left(fileName, len(fileName) - len(basename(fileName)))                            
                                                
                        if self.combine == True:
                            self.value.emit('\nSaving File to Database...')
                            engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                            dfExcel.uploadCSV(basename(fileName), engine, 50000)

                        else:
                            self.value.emit('\nSaving File to Database...')
                            engine = "DRIVER={ODBC Driver 13 for SQL Server};SERVER="+self.server+";DATABASE="+self.database+";Trusted_Connection=yes"
                            dfExcel.uploadCSV('pass', engine, 50000)
                    
                    else: 
                        
                        if self.combine == True:
                            pass
                        elif self.combine == False and self.saveFile != '': 
                            csv_files = [i for i in glob.glob('*.{}'.format('gz'))]
                            for i in csv_files:
                                move(i, filePath + i)
                        
                        
                        if '.' in basename(self.saveFile):
                            fileName = os.path.splitext(self.saveFile)[0] 
                            filePath = left(fileName, len(fileName) - len(basename(fileName)))
                        else: 
                            fileName = self.saveFile
                            filePath = left(fileName, len(fileName) - len(basename(fileName)))
                        
                        if self.combine == True:
                            
                            if self.zipped == True:
                                self.value.emit('\nSaving File to csv.gz...')
                                dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                            else:
                                self.value.emit('\nSaving File to csv...')
                                dfExcel.combineCSV(fileName, filePath, zipped = self.zipped, delim = self.delimOut)
                        
                        else:
                            if self.zipped == True:
                                csv_files = [i for i in glob.glob('*.{}'.format('gz'))]
                                for i in csv_files:
                                    move(i, filePath + i)
                            else:
                                csv_files = [i for i in glob.glob('*.{}'.format('csv'))]
                                for i in csv_files:
                                    move(i, filePath + i)
            
            writeend= time.time()
            self.value.emit('\nSave Complete.')
            self.value.emit('\nWrite time: ' + str(round(writeend - writestart,2)))
            self.value.emit("\nFormatting Complete.")
            end = time.time()
            
            csv_files = [i for i in glob.glob('*.{}'.format('csv'))]
            
            for i in csv_files:
                os.remove(i)
                
            flat_files =[]
            non_flat = []
            xls = []
            count = 0
            sheets = {}
            
            self.value.emit("\nTotal Time Taken: " + str(round(end - start, 2)) + " Seconds.") 
            
            self.runCont.emit() 
            
            self.timerValue.emit(35)
            self.endValue.emit(1)
            gc.collect()
            
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            self.value.emit(str(e))
            self.endValue.emit(1)   
            self.runCont.emit()


if __name__ == '__main__':   
    print('Program Start')
    freeze_support()
    app = QApplication(sys.argv)  
    app.setStyle('Fusion')
    ex = App()
    sys.exit(app.exec_())      